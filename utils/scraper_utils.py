import os
from bs4 import BeautifulSoup
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from pathlib import Path

from crawl4ai import (
    AsyncWebCrawler,
    BrowserConfig,
    CacheMode,
    CrawlerRunConfig,
)

GREEN_CLASS = "kwOsxd"
RED_CLASS   = "bgICll"


def get_browser_config() -> BrowserConfig:
    return BrowserConfig(
        browser_type="chromium",
        headless=False,
        verbose=True,
    )


def get_signed_value(cell):
    p = cell.find("p")
    if not p:
        return cell.get_text(strip=True)
    classes = p.get("class", [])
    value = p.get_text(strip=True)
    if RED_CLASS in classes:
        return f"-{value}"
    elif GREEN_CLASS in classes:
        return f"+{value}"
    return value


def get_color(cell) -> str:
    p = cell.find("p")
    if not p:
        return "neutral"
    classes = p.get("class", [])
    if RED_CLASS in classes:
        return "red"
    elif GREEN_CLASS in classes:
        return "green"
    return "neutral"


def parse_investments_from_html(html: str):
    soup = BeautifulSoup(html, "html.parser")
    rows = soup.select("tr.ant-table-row")

    investments = []
    for row in rows:
        cells = row.select("td.ant-table-cell")
        if len(cells) < 8:
            continue
        investments.append({
            "Asset":                   cells[0].get_text(strip=True),
            "Asset Class":             cells[1].get_text(strip=True),
            "Units Owned":             cells[2].get_text(strip=True),
            "Cost Per Unit":           cells[3].get_text(strip=True),
            "Current Price":           cells[4].get_text(strip=True),
            "Market Value":            cells[5].get_text(strip=True),
            "Daily Change":            get_signed_value(cells[6]),
            "Daily Change Color":      get_color(cells[6]),
            "Unrealized Return":       get_signed_value(cells[7]),
            "Unrealized Return Color": get_color(cells[7]),
        })

    return investments


async def fetch_investments(url: str) -> List[dict]:
    session_id = "thndr_session"

    async with AsyncWebCrawler(config=get_browser_config()) as crawler:

        print("\n⏳ Opening browser...")
        print("👉 Please log in. Once you can see your investments, press ENTER here.\n")

        await crawler.arun(
            url=url,
            config=CrawlerRunConfig(
                cache_mode=CacheMode.BYPASS,
                session_id=session_id,
                page_timeout=180000,
                delay_before_return_html=2.0,
            ),
        )

        input("✅ Press ENTER when you can see your investments in the browser...")

        print("\n🔍 Capturing page data...")
        result = await crawler.arun(
            url=url,
            config=CrawlerRunConfig(
                cache_mode=CacheMode.BYPASS,
                session_id=session_id,
                delay_before_return_html=3.0,
                page_timeout=30000,
            ),
        )

    if not result.success:
        print(f"❌ Error: {result.error_message}")
        return []

    investments = parse_investments_from_html(result.cleaned_html or "")
    print(f"✅ Found {len(investments)} investments.")
    return investments


def get_dated_filepath(folder: str) -> str:
    """Auto-increments filename so existing files are never overwritten."""
    today = date.today().strftime("%Y-%m-%d")
    counter = 1
    while True:
        filepath = Path(folder) / f"investments_{today}_{counter}.xlsx"
        if not filepath.exists():
            return str(filepath)
        counter += 1


def save_to_excel(investments: List[dict], folder: str):
    if not investments:
        print("Nothing to save.")
        return

    # Make sure the folder exists (works on any PC)
    Path(folder).mkdir(parents=True, exist_ok=True)

    filepath = get_dated_filepath(folder)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Investments"

    header_fill = PatternFill("solid", fgColor="1F2D3D")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill    = PatternFill("solid", fgColor="FFDADA")
    green_fill  = PatternFill("solid", fgColor="DAF5DA")
    red_font    = Font(color="C0392B", bold=True)
    green_font  = Font(color="1E8449", bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = ["Asset", "Asset Class", "Units Owned", "Cost Per Unit",
               "Current Price", "Market Value", "Daily Change", "Unrealized Return"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = thin_border

    for row_idx, inv in enumerate(investments, start=2):
        values = [
            inv["Asset"],
            inv["Asset Class"],
            inv["Units Owned"],
            inv["Cost Per Unit"],
            inv["Current Price"],
            inv["Market Value"],
            inv["Daily Change"],
            inv["Unrealized Return"],
        ]
        colors = [None, None, None, None, None, None,
                  inv["Daily Change Color"],
                  inv["Unrealized Return Color"]]

        for col_idx, (val, color) in enumerate(zip(values, colors), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center
            cell.border    = thin_border

            if color == "red":
                cell.fill = red_fill
                cell.font = red_font
            elif color == "green":
                cell.fill = green_fill
                cell.font = green_font
            elif row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7F9FC")

    col_widths = [10, 14, 13, 14, 14, 14, 22, 22]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"💾 Saved to {filepath}")